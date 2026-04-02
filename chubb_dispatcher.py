#!/usr/bin/env python3
"""
The CHUBB Dispatcher
EFT / Payment Transmittal CSV formatter.

Double-click  → GUI launcher (supports drag-and-drop onto input field)
CLI:  CHUBB_Dispatcher.exe <file_or_folder> [--output <folder>]
"""

import csv
import sys
import os
import argparse
import traceback
from pathlib import Path
from datetime import datetime

# ── PyInstaller fix for tkinterdnd2 ──────────────────────────────────────────
# tkinterdnd2._require() resolves the tkdnd DLL path using:
#   os.path.dirname(__file__)  →  'tkdnd'  →  'win-x64'
# Inside a --onefile exe, __file__ points into the PYZ archive (not a real
# directory on disk), so that path never exists and DnD silently fails.
# Fix: patch TkinterDnD.__file__ to point at the real sys._MEIPASS location
# before the Tk() constructor calls _require().
if getattr(sys, 'frozen', False):
    import tkinterdnd2.TkinterDnD as _tkdnd_mod
    _tkdnd_mod.__file__ = os.path.join(
        sys._MEIPASS, 'tkinterdnd2', 'TkinterDnD.py'
    )

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS_TO_DELETE = {
    'CLAIMANT-NAME',
    'PATIENT-#',
    'DATES-OF-SVC.-OR-PYMT',
    'ORIG-BILL-AMT',
    'PAID TO',
    'EFT REF NUMBER',
    'CHUBB OFFICE',
    'AGENCY-CLAIM-#',
}

PAID_AMT_FORMAT = '$#,##0.00_);[Red]($#,##0.00)'
DATE_FORMAT     = 'm/d/yyyy'
FOOTER_MARKER   = 'TOTAL DEPOSIT AMOUNT:'

# ─────────────────────────────────────────────────────────────────────────────
# CSV helpers
# ─────────────────────────────────────────────────────────────────────────────

def read_csv(file_path: Path) -> list:
    with open(file_path, newline='', encoding='utf-8-sig') as fh:
        return list(csv.reader(fh))


def find_footer_index(rows: list):
    for i, row in enumerate(rows):
        if any(FOOTER_MARKER in cell for cell in row):
            return i
    return None


def is_blank(row: list) -> bool:
    return all(cell == '' for cell in row)


# ─────────────────────────────────────────────────────────────────────────────
# Processing  (Steps 1 – 3b)
# ─────────────────────────────────────────────────────────────────────────────

def process_rows(rows: list) -> list:
    if not rows:
        raise ValueError("File is empty.")

    header = [c.strip() for c in rows[0]]
    header = ['INVOICE' if h == 'INVOICE-NUMBER' else h for h in header]

    keep       = [i for i, h in enumerate(header) if h not in COLUMNS_TO_DELETE]
    new_header = [header[i] for i in keep]

    result = [new_header]
    for raw in rows[1:]:
        padded  = (list(raw) + [''] * len(header))[:len(header)]
        trimmed = [c.strip() for c in padded]
        result.append([trimmed[i] for i in keep])

    fi = find_footer_index(result)
    if fi is not None and fi > 0:
        scan = fi - 1
        while scan > 0 and is_blank(result[scan]):
            scan -= 1
        blank_count = fi - scan - 1
        if blank_count == 0:
            result.insert(fi, [''] * len(new_header))
        elif blank_count > 1:
            del result[scan + 1 : fi - 1]

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Type coercion
# ─────────────────────────────────────────────────────────────────────────────

def try_parse_date(value: str):
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return value


def try_parse_amount(value: str):
    if not value:
        return value
    cleaned = value.replace('$', '').replace(',', '').strip()
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned)
    except ValueError:
        return value


# ─────────────────────────────────────────────────────────────────────────────
# XLSX writer  (Steps 4 – 6)
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(rows: list, output_path: Path) -> int:
    if not rows:
        raise ValueError("No data rows to write.")

    wb = openpyxl.Workbook()
    ws = wb.active

    header = rows[0]
    n_cols = len(header)
    fi     = find_footer_index(rows)

    invoice_count = sum(
        1 for i, row in enumerate(rows)
        if i > 0 and i != fi and not is_blank(row)
    )

    for r, row in enumerate(rows, start=1):
        is_hdr = (r == 1)
        is_ftr = (r - 1 == fi)
        is_blk = is_blank(row)
        is_dat = not is_hdr and not is_ftr and not is_blk

        for c, raw in enumerate(row, start=1):
            col = header[c - 1] if c <= n_cols else ''
            val = raw
            if raw and (is_dat or is_ftr):
                if col == 'PAID-AMT':
                    val = try_parse_amount(raw)
                elif col in ('STATEMENT DATE', 'DEPOSIT DATE') and is_dat:
                    val = try_parse_date(raw)
            ws.cell(row=r, column=c, value=val)

    total_rows = ws.max_row
    thin   = Side(style='thin')
    thick  = Side(style='thick')
    medium = Side(style='medium')

    for r in range(1, total_rows + 1):
        src    = rows[r - 1] if r - 1 < len(rows) else []
        is_ftr = (r - 1 == fi)
        is_blk = is_blank(src)
        is_dat = r > 1 and not is_ftr and not is_blk

        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            col  = header[c - 1] if c <= n_cols else ''

            cell.font      = Font(name='Calibri', size=18, bold=is_ftr)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if col == 'PAID-AMT' and (is_dat or is_ftr):
                cell.number_format = PAID_AMT_FORMAT
            elif col in ('STATEMENT DATE', 'DEPOSIT DATE') and is_dat:
                cell.number_format = DATE_FORMAT

            top_s   = thick if r == 1          else thin
            bot_s   = thick if r == total_rows else thin
            left_s  = thick if c == 1          else thin
            right_s = thick if c == n_cols     else thin

            if is_ftr:
                top_s  = medium
                bot_s  = medium
                if c == 1:      left_s  = medium
                if c == n_cols: right_s = medium

            cell.border = Border(top=top_s, bottom=bot_s,
                                 left=left_s, right=right_s)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        max_len = max(
            (len(str(ws.cell(row=r, column=c).value))
             for r in range(1, total_rows + 1)
             if ws.cell(row=r, column=c).value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = max(max_len * 1.8 + 2, 12)

    for r in range(1, total_rows + 1):
        ws.row_dimensions[r].height = 32

    wb.save(output_path)
    return invoice_count


# ─────────────────────────────────────────────────────────────────────────────
# Batch runner
# ─────────────────────────────────────────────────────────────────────────────

def process_file(input_path: Path, output_path: Path) -> int:
    return write_xlsx(process_rows(read_csv(input_path)), output_path)


def collect_csv_files(input_paths: list) -> list:
    files, seen = [], set()
    for p in input_paths:
        p = Path(p)
        candidates = (
            sorted(p.glob('*.csv')) + sorted(p.glob('*.CSV'))
            if p.is_dir() else ([p] if p.is_file() else [])
        )
        for f in candidates:
            key = f.resolve()
            if key not in seen:
                seen.add(key)
                files.append(f)
    return files


def run_batch(input_paths: list, output_dir=None, log=print):
    files = collect_csv_files(input_paths)
    if not files:
        log("No CSV files found.")
        return [], []

    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    results, errors = [], []
    log(f"Processing {len(files)} file(s)...\n")

    for fp in files:
        dest     = output_dir if output_dir else fp.parent
        out_path = dest / (fp.stem + '.xlsx')
        try:
            count = process_file(fp, out_path)
            results.append((fp.name, count, out_path))
            log(f"  [OK]     {fp.name}  ->  {out_path.name}  ({count} invoice(s))")
        except Exception as exc:
            errors.append((fp.name, str(exc)))
            log(f"  [ERROR]  {fp.name}: {exc}")
            log(traceback.format_exc())

    total = sum(c for _, c, _ in results)
    log("")
    log("=" * 58)
    log("  SUMMARY")
    log("=" * 58)
    for fname, count, op in results:
        log(f"  {fname}")
        log(f"    -> {op}  ({count} invoice(s))")
    if errors:
        log(f"\n  ERRORS ({len(errors)}):")
        for fname, reason in errors:
            log(f"    {fname}: {reason}")
    log(f"\n  Files processed : {len(results)}")
    log(f"  Total invoices  : {total}")
    if errors:
        log(f"  Files with errors: {len(errors)}")
    log("=" * 58)

    return results, errors


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def run_cli():
    parser = argparse.ArgumentParser(
        prog='CHUBB_Dispatcher',
        description='The CHUBB Dispatcher — EFT transmittal CSV formatter',
    )
    parser.add_argument('input', help='CSV file or folder containing CSV files')
    parser.add_argument('--output', '-o', default=None,
                        help='Output folder (default: same directory as input)')
    args = parser.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        print(f'Error: "{inp}" does not exist.')
        sys.exit(1)

    print("\nThe CHUBB Dispatcher\n")
    run_batch([inp], output_dir=args.output)


# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────

def run_gui():
    import tkinter as tk
    from tkinter import filedialog, scrolledtext
    from tkinterdnd2 import TkinterDnD, DND_FILES
    import threading

    BLUE  = '#003087'
    LBLUE = '#d0e4f7'
    BG    = '#f4f7fb'
    WHITE = '#ffffff'

    class App(TkinterDnD.Tk):
        def __init__(self):
            super().__init__()
            self.title('The CHUBB Dispatcher')
            self.configure(bg=BG)
            self.minsize(680, 520)
            self.resizable(True, True)
            self._selected: list = []
            self._build()
            self.update_idletasks()
            w, h = self.winfo_width(), self.winfo_height()
            x = (self.winfo_screenwidth()  - w) // 2
            y = (self.winfo_screenheight() - h) // 2
            self.geometry(f'+{x}+{y}')

        # ── Build UI ──────────────────────────────────────────────────────

        def _build(self):
            # Header
            hdr = tk.Frame(self, bg=BLUE)
            hdr.pack(fill='x')
            tk.Label(hdr, text='THE CHUBB DISPATCHER',
                     bg=BLUE, fg=WHITE,
                     font=('Calibri', 22, 'bold'), pady=10).pack()
            tk.Label(hdr, text='EFT / Payment Transmittal Formatter',
                     bg=BLUE, fg=LBLUE,
                     font=('Calibri', 10)).pack(pady=(0, 8))

            # Body
            body = tk.Frame(self, bg=BG, padx=18, pady=14)
            body.pack(fill='both', expand=True)
            body.columnconfigure(0, weight=1)

            # Input label
            tk.Label(body,
                     text='Input — drag a file/folder here, or browse:',
                     bg=BG, font=('Calibri', 11, 'bold')).grid(
                         row=0, column=0, sticky='w', pady=(0, 4))

            # Input row
            inp_frame = tk.Frame(body, bg=BG)
            inp_frame.grid(row=1, column=0, sticky='ew', pady=(0, 10))
            inp_frame.columnconfigure(0, weight=1)

            self._inp_var = tk.StringVar()
            self._inp_entry = tk.Entry(
                inp_frame, textvariable=self._inp_var,
                font=('Calibri', 10), state='readonly',
                readonlybackground=WHITE, relief='solid', bd=1,
            )
            self._inp_entry.grid(row=0, column=0, sticky='ew',
                                 padx=(0, 6), ipady=3)

            # Register drag-and-drop on the entry field
            self._inp_entry.drop_target_register(DND_FILES)
            self._inp_entry.dnd_bind('<<Drop>>',      self._on_drop)
            self._inp_entry.dnd_bind('<<DragEnter>>', self._on_drag_enter)
            self._inp_entry.dnd_bind('<<DragLeave>>', self._on_drag_leave)

            btn_f = tk.Frame(inp_frame, bg=BG)
            btn_f.grid(row=0, column=1)
            self._btn(btn_f, 'File(s)...', self._pick_files).pack(
                side='left', padx=(0, 4))
            self._btn(btn_f, 'Folder...', self._pick_folder).pack(side='left')

            # Output label + row
            tk.Label(body,
                     text='Output folder  (blank = save next to input):',
                     bg=BG, font=('Calibri', 11, 'bold')).grid(
                         row=2, column=0, sticky='w', pady=(0, 4))

            out_frame = tk.Frame(body, bg=BG)
            out_frame.grid(row=3, column=0, sticky='ew', pady=(0, 14))
            out_frame.columnconfigure(0, weight=1)

            self._out_var = tk.StringVar()
            tk.Entry(out_frame, textvariable=self._out_var,
                     font=('Calibri', 10), relief='solid', bd=1).grid(
                         row=0, column=0, sticky='ew', padx=(0, 6), ipady=3)
            self._btn(out_frame, 'Browse...', self._pick_output).grid(
                row=0, column=1)

            # Process button
            self._proc_btn = tk.Button(
                body, text='Process Files',
                command=self._on_process,
                bg=BLUE, fg=WHITE,
                font=('Calibri', 14, 'bold'),
                relief='flat', padx=28, pady=10,
                activebackground='#00215e',
                activeforeground=WHITE,
                cursor='hand2',
            )
            self._proc_btn.grid(row=4, column=0, pady=(0, 12))

            # Log area
            tk.Label(body, text='Output log:',
                     bg=BG, font=('Calibri', 10, 'bold')).grid(
                         row=5, column=0, sticky='w')
            self._log = scrolledtext.ScrolledText(
                body, font=('Consolas', 9),
                height=13, state='disabled', wrap='word',
                bg='#1e1e1e', fg='#d4d4d4',
                insertbackground=WHITE, relief='flat',
            )
            self._log.grid(row=6, column=0, sticky='nsew', pady=(3, 0))
            body.rowconfigure(6, weight=1)

            self._log.tag_config('ok',  foreground='#6fcf97')
            self._log.tag_config('err', foreground='#eb5757')
            self._log.tag_config('hdr', foreground='#56ccf2',
                                 font=('Consolas', 9, 'bold'))

        # ── Helpers ───────────────────────────────────────────────────────

        @staticmethod
        def _btn(parent, text, cmd):
            return tk.Button(
                parent, text=text, command=cmd,
                font=('Calibri', 10), relief='solid', bd=1,
                padx=8, pady=3, cursor='hand2',
                bg=WHITE, activebackground=LBLUE,
            )

        def _log_write(self, msg, tag=None):
            self._log.config(state='normal')
            if tag:
                self._log.insert('end', msg + '\n', tag)
            else:
                self._log.insert('end', msg + '\n')
            self._log.see('end')
            self._log.config(state='disabled')

        def _log_clear(self):
            self._log.config(state='normal')
            self._log.delete('1.0', 'end')
            self._log.config(state='disabled')

        def _set_selected(self, paths: list):
            self._selected = paths
            if len(paths) == 1:
                self._inp_var.set(paths[0])
            else:
                first = Path(paths[0]).name
                self._inp_var.set(
                    f'{len(paths)} files selected  (first: {first})')

        # ── Drag-and-drop ─────────────────────────────────────────────────

        def _on_drop(self, event):
            paths = list(self.tk.splitlist(event.data))
            if paths:
                self._set_selected(paths)
            self._inp_entry.config(readonlybackground=WHITE)

        def _on_drag_enter(self, event):
            self._inp_entry.config(readonlybackground=LBLUE)

        def _on_drag_leave(self, event):
            self._inp_entry.config(readonlybackground=WHITE)

        # ── Browse ────────────────────────────────────────────────────────

        def _pick_files(self):
            paths = filedialog.askopenfilenames(
                title='Select CSV file(s)',
                filetypes=[('CSV files', '*.csv *.CSV'), ('All files', '*.*')],
            )
            if paths:
                self._set_selected(list(paths))

        def _pick_folder(self):
            path = filedialog.askdirectory(
                title='Select folder containing CSV files')
            if path:
                self._set_selected([path])

        def _pick_output(self):
            path = filedialog.askdirectory(title='Select output folder')
            if path:
                self._out_var.set(path)

        # ── Processing ────────────────────────────────────────────────────

        def _on_process(self):
            if not self._selected:
                from tkinter import messagebox
                messagebox.showerror(
                    'No Input',
                    'Please select or drag in a CSV file or folder first.',
                    parent=self,
                )
                return
            self._proc_btn.config(state='disabled', text='Processing...')
            self._log_clear()
            threading.Thread(target=self._worker, daemon=True).start()

        def _worker(self):
            out = self._out_var.get().strip() or None

            def log(msg):
                tag = None
                s = msg.strip()
                if s.startswith('[OK]'):       tag = 'ok'
                elif s.startswith('[ERROR]'):  tag = 'err'
                elif s.startswith('===') or s == 'SUMMARY': tag = 'hdr'
                self.after(0, self._log_write, msg, tag)

            results, errors = run_batch(self._selected, output_dir=out, log=log)

            def finish():
                self._proc_btn.config(state='normal', text='Process Files')
                total = sum(c for _, c, _ in results)
                if results and not errors:
                    from tkinter import messagebox
                    messagebox.showinfo(
                        'Done',
                        f'{len(results)} file(s) processed successfully.\n'
                        f'{total} total invoice(s).',
                        parent=self,
                    )
                elif errors:
                    from tkinter import messagebox
                    messagebox.showwarning(
                        'Completed with errors',
                        f'{len(results)} file(s) OK, {len(errors)} error(s).\n'
                        'See the output log for details.',
                        parent=self,
                    )

            self.after(0, finish)

    App().mainloop()


# ─────────────────────────────────────────────────────────────────────────────
# Entry
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) > 1:
        run_cli()
    else:
        run_gui()
